import os
import requests
import json
import openpyxl
import argparse

# Parse command line arguments
parser = argparse.ArgumentParser(description="Fetch and export GitHub repositories")
parser.add_argument("organization", help="GitHub organization name")
args = parser.parse_args()

# Get the access token from the environment variable
access_token = os.getenv("GITHUB_OAUTH_TOKEN")

# Set the GitHub organization name from the command line argument
organization = args.organization

# Create a headers dictionary with the access token
headers = {"Authorization": f"Bearer {access_token}"}

# Create a list to store the repository data
repository_data = []

# Set the initial page number
page = 1

# Fetch repositories until there are no more pages
while True:
    # Fetch the list of repositories from the organization
    response = requests.get(f"https://api.github.com/orgs/{organization}/repos?page={page}", headers=headers)
    if response.status_code == 200:
        repositories = response.json()

        # If there are no more repositories, break the loop
        if not repositories:
            break

        # Loop through the list of repositories and export the data
        for repository in repositories:
            response = requests.get(repository["url"], headers=headers)
            if response.status_code == 200:
                repo_data = response.json()
                repo_data["archived"] = repository["archived"]  # Add the archived status
                repository_data.append(repo_data)
            else:
                print(f"Error exporting repository {repository['url']}: {response.status_code}")

        # Increment the page number
        page += 1
    else:
        print(f"Error fetching repositories from organization {organization}: {response.status_code}")
        break

# Export the repository data to an Excel file
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Set the column headers
worksheet["A1"] = "ID"
worksheet["B1"] = "Name"
worksheet["C1"] = "Description"
worksheet["D1"] = "URL"
worksheet["E1"] = "Created At"
worksheet["F1"] = "Updated At"
worksheet["G1"] = "Language"
worksheet["H1"] = "Stargazers Count"
worksheet["I1"] = "Forks Count"
worksheet["J1"] = "Open Issues Count"
worksheet["K1"] = "Archived"  # Add the column header for archived status

# Add data to the worksheet
for row, repo in enumerate(repository_data, start=2):
    worksheet.cell(row=row, column=1).value = repo["id"]
    worksheet.cell(row=row, column=2).value = repo["name"]
    worksheet.cell(row=row, column=3).value = repo["description"]
    worksheet.cell(row=row, column=4).value = repo["html_url"]
    worksheet.cell(row=row, column=5).value = repo["created_at"]
    worksheet.cell(row=row, column=6).value = repo["updated_at"]
    worksheet.cell(row=row, column=7).value = repo["language"]
    worksheet.cell(row=row, column=8).value = repo["stargazers_count"]
    worksheet.cell(row=row, column=9).value = repo["forks_count"]
    worksheet.cell(row=row, column=10).value = repo["open_issues_count"]
    worksheet.cell(row=row, column=11).value = repo["archived"]  # Add the archived status

# Save the workbook to a file
workbook.save("output.xlsx")

# Write the repository data to a JSON file
with open("output.json", "w") as file:
    json.dump(repository_data, file, indent=4)